from datetime import datetime
import calendar
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

def process(file_input, file_output):


    # ---------------- Row Missed Count ----------------

    # Read header info
    header = pd.read_excel(file_input, header=None, nrows=4, usecols="L")

    # Extract community and date
    community = header.iloc[0, 0]
    excel_date = header.iloc[1, 0]

    split_date = excel_date.split()

    month_str = split_date[0]
    month_int = datetime.strptime(month_str, '%B').month
    year_int = int(split_date[1])

    _, num_days = calendar.monthrange(year_int, month_int)

    valid_days = set(range(1, num_days + 1))

    # Create datetime column mapping
    rename_dict = {
        day: datetime(year_int, month_int, day)
        for day in range(1, num_days + 1)
    }

    # Read Charting info and convert day and date columns to datetime
    df = pd.read_excel(file_input, skiprows=4)
    df.rename(columns=rename_dict, inplace=True)

    df = df[
        [col for col in df.columns
        if not (isinstance(col, datetime) and col.day not in valid_days)]
    ].copy()
    
    df['Alert Charting Start Date'] = pd.to_datetime(df['Alert Charting Start Date'])
    df['Alert Charting End Date'] = pd.to_datetime(df['Alert Charting End Date'])

    # Remove empty rows
    df = df[df['Resident Name'].notna()].reset_index(drop=True)

    # Count missed by row (resident)
    def count_row_missed(row):

        missed = 0
        start, end = row['Alert Charting Start Date'], row['Alert Charting End Date']

        # Fail if missing start or end charting dates
        if pd.isna(start) or pd.isna(end):
            return 0

        # If start date but no end date, set end date to the last day of the month
        if pd.isna(end):
            end = datetime(year_int, month_int, num_days)


        for col in df.columns:
            if isinstance(col, datetime):
                if start <= col <= end:
                    if pd.isna(row[col]) or str(row[col]).strip().upper() != 'X': # X Represents a successful chart
                        missed += 1

        return missed

    df['Missed'] = df.apply(count_row_missed, axis=1)

    # Convert start/end back to date part only
    df['Alert Charting Start Date'] = df['Alert Charting Start Date'].dt.date
    df['Alert Charting End Date'] = df['Alert Charting End Date'].dt.date

    # Visually move End Date next to Start Date
    cols = list(df.columns)
    cols.remove('Alert Charting End Date')
    start_idx = cols.index('Alert Charting Start Date')
    cols.insert(start_idx + 1, 'Alert Charting End Date')
    df = df[cols]

    # Convert datetime column headers back to day integers
    new_cols = []
    for col in df.columns:
        if isinstance(col, datetime):
            new_cols.append(col.day)
        else:
            new_cols.append(col)

    df.columns = new_cols

    # Save file to output
    df.to_excel(file_output, index=False)


    # ---------------- Highlighting ----------------

    # Read the excel file that was just saved
    wb = load_workbook(file_output)
    ws = wb.active

    # General formatting
    center_align = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align

    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    for cell in ws[1]:
        cell.fill = header_fill

    for cell in ws[1]:  # first row
        cell.font = bold_font
        cell.border = thin_border

    # Highlight days that should have been marked with an 'X' but were not, these are the missed days
    highlight = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Map day columns again (now integers)
    col_map = {}
    for cell in ws[1]:
        if isinstance(cell.value, int) and cell.value in valid_days:
            col_map[cell.value] = cell.column

    # Find Missed column index
    missed_col_idx = None
    for cell in ws[1]:
        if cell.value == 'Missed':
            missed_col_idx = cell.column
            break

    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if missed_col_idx:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=missed_col_idx, max_col=missed_col_idx):
            for cell in row:
                cell.font = bold_font
                cell.border = thin_border

    # Highlight logic
    for r_idx, row in enumerate(df.itertuples(), start=2):

        start = df.at[r_idx-2, 'Alert Charting Start Date']
        end = df.at[r_idx-2, 'Alert Charting End Date']

        if pd.isna(start):
            continue

        if pd.isna(end):
            end = datetime(year_int, month_int, num_days).date()

        # Highlight missed day cells
        for day, col_idx in col_map.items():

            date_obj = datetime(year_int, month_int, day).date()

            if start <= date_obj <= end:
                cell = ws.cell(row=r_idx, column=col_idx)
                val = str(cell.value).strip().upper() if cell.value is not None else ""

                if val != 'X':
                    cell.fill = highlight

        # Highlight Missed column if > 0
        if missed_col_idx:
            missed_val = df.at[r_idx-2, 'Missed']
            if pd.notna(missed_val) and missed_val > 0:
                ws.cell(row=r_idx, column=missed_col_idx).fill = highlight

    ws.insert_rows(1, amount=3)
    ws.cell(row=1, column=1, value='Missed Alert Charting Monthly')
    ws.cell(row=2, column=1, value=community)
    ws.cell(row=3, column=1, value=excel_date)


    # ---------------- Column Missed Count ----------------

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value='DAILY TOTAL MISSED')

    # count highlighted cells per date column
    for day, col_idx in col_map.items():

        count = 0

        for r in range(4, ws.max_row):  # skip title (3 rows) + header row
            cell = ws.cell(row=r, column=col_idx)

            if cell.fill == highlight:
                count += 1

        ws.cell(row=total_row, column=col_idx, value=count)

    # Missed column total (same idea)
    if missed_col_idx:

        total = 0

        for r in range(4, ws.max_row):  # skip headers + title rows
            cell = ws.cell(row=r, column=missed_col_idx)

            if isinstance(cell.value, (int, float)):
                total += cell.value

        ws.cell(row=total_row, column=missed_col_idx, value=total)

    # style total row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Highlight Total row if > 0

    for cell in ws[total_row]:
        if isinstance(cell.value, (int, float)) and cell.value > 0:
            cell.fill = highlight

    wb.save(file_output)
